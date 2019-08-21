from xlrd import open_workbook
from xlutils.copy import copy
import copy
import xlutils
import re, time
import tkinter as tk
import openpyxl
from openpyxl import load_workbook
from tkinter import *
import tkinter.filedialog
import tkinter.messagebox
from tkinter import ttk


def main():
    root = tk.Tk()
    # fix the root window size
    root.minsize(300, 120)
    root.maxsize(500, 120)  # 这里主要是控制窗口的大小，让窗口大小不能改变
    root.title('身份证计算年龄工具')  # 设置主窗口的标题
    path = StringVar()
    lb = Label(root, text="目标路径:")
    lb.grid(row=0, column=0)
    entry = Entry(root, textvariable=path)
    entry.grid(row=0, column=1, columnspan=3)
    ljBt = Button(root, text="路径选择", command=lambda: xz(path))
    ljBt.grid(row=0, column=4)
    # 第二排
    # 选择表格页
    v = IntVar()
    v.set(1)
    selectSheetTitle = Label(root, text="请选择表格页（第几页）:")
    selectSheetTitle.grid(row=1, column=0)
    sheet1 = Radiobutton(root, text='1', variable=v, value='1')
    sheet2 = Radiobutton(root, text='2', variable=v, value='2')
    sheet3 = Radiobutton(root, text='3', variable=v, value='3')
    sheet1.grid(row=1, column=1)
    sheet2.grid(row=1, column=2)
    sheet3.grid(row=1, column=3)
    # 第三行选择第几列来解析
    selectSFZTitle = Label(root, text="请选择第几列:")
    selectSFZTitle.grid(row=2, column=0)
    number = tk.StringVar()
    numberChosen = ttk.Combobox(root, textvariable=number)
    numberChosen['values'] = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10)  # 设置下拉列表的值
    numberChosen.grid(row=2, column=2)  # 设置其在界面中出现的位置  column代表列   row 代表行
    numberChosen.current(3)
    # 第四行，操作行
    # 计算
    c_age = Button(root, text="确定", command=lambda: cAge(v.get(), entry.get(),numberChosen.get()))
    c_age.grid(row=3, column=2)
    root.mainloop()


def date_style_transfomation(date, format_string1="%Y%m%d", format_string2="%Y-%m-%d"):
    time_array = time.strptime(date, format_string1)
    str_date = time.strftime(format_string2, time_array)
    time_stamp = int(time.mktime(time_array))
    now_time_stamp = int(time.mktime(time.localtime()))
    if (time_stamp < now_time_stamp):
        return 1;
    else:
        return 0;


def excel2003xls(name, sheetNum,numberChosen):
    data = open_workbook(name)
    table1 = data.sheet_by_index(sheetNum)
    newDate = copy(data)
    table = newDate.get_sheet(0)
    saveName = name.split(".")[0] + "1." + name.split(".")[1]
    rowNum = table1.nrows
    colNum = table1.ncols
    if rowNum == 0 or colNum == 0:
        return;
    # 获取所有单元格的内容
    # 身份证 正则
    pattern = re.compile(
        r'(^[1-9]\d{5}(18|19|([23]\d))\d{2}((0[1-9])|(10|11|12))(([0-2][1-9])|10|20|30|31)\d{3}[0-9Xx])|([1−9]\d5\d2((0[1−9])|(10|11|12))(([0−2][1−9])|10|20|30|31)\d2[0−9Xx])')
    for i in range(rowNum):
        sfz = table1.cell_value(i, int(numberChosen)-1);
        if re.findall(pattern, str(sfz)):
            y = int(sfz[6:10])
            now_y = int(time.strftime('%Y', time.localtime()));
            age = now_y - y
            isadd = date_style_transfomation(time.strftime('%Y', time.localtime()) + sfz[10:14])
            if isadd == 1:
                age = age - 1
            table.write(i, int(colNum) + 1, age)
    newDate.save(saveName)
    tk.messagebox.showinfo(title='success', message='执行完毕')


def excel2007xlsx(name, sheetNum,numberChosen):
    worksheet = load_workbook(name)  # 打开excel文件
    sheet_names = worksheet.sheetnames
    sheet1 = worksheet[sheet_names[sheetNum]]
    saveName = name.split(".")[0] + "1." + name.split(".")[1]
    rowNum = sheet1.max_row
    colNum = sheet1.max_column
    if rowNum == 0 or colNum == 0:
        return;
    # 获取所有单元格的内容
    # 身份证 正则
    pattern = re.compile(
        r'(^[1-9]\d{5}(18|19|([23]\d))\d{2}((0[1-9])|(10|11|12))(([0-2][1-9])|10|20|30|31)\d{3}[0-9Xx])|([1−9]\d5\d2((0[1−9])|(10|11|12))(([0−2][1−9])|10|20|30|31)\d2[0−9Xx])')
    for i in range(1, rowNum + 1):
        sfz = sheet1.cell(row=i, column=int(numberChosen)).value;
        if re.findall(pattern, str(sfz)):
            y = int(sfz[6:10])
            now_y = int(time.strftime('%Y', time.localtime()));
            age = now_y - y
            isadd = date_style_transfomation(time.strftime('%Y', time.localtime()) + sfz[10:14])
            if isadd == 1:
                age = age - 1
            sheet1.cell(row=i, column=int(colNum) + 1).value = age
    worksheet.save(saveName)
    tk.messagebox.showinfo(title='success', message='执行完毕')


def xz(entry):
    filename = tkinter.filedialog.askopenfilename(
        filetypes=(("Excel 2003 files", "*.xls*"), ("Excel 2007 file", "*.xlsx*")))
    if filename != '':
        # if len(entry.get())>0:
        #     entry.delete('1.0', 'end')
        entry.set(filename)
    else:
        tk.messagebox.showinfo(title='error', message='请选择正确的文件')


def cAge(sheetNum, path,numberChosen):
    if path.split("/")[-1].split(".")[1] == 'xlsx':
        excel2007xlsx(path, int(sheetNum) - 1,numberChosen)
    if path.split("/")[-1].split(".")[1] == 'xls':
        excel2003xls(path, int(sheetNum) - 1,numberChosen)


if __name__ == '__main__':
    main()
