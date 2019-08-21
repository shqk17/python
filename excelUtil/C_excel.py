from openpyxl import load_workbook

from excelUtil import commonUtils


def main():
    excel2007xlsx("./红黄蓝亲子园客户导入模板.xlsx", 0, 3)

def excel2007xlsx(name, sheetNum,numberChosen):
    worksheet = load_workbook(name)  # 打开excel文件
    sheet_names = worksheet.sheetnames
    sheet1 = worksheet[sheet_names[sheetNum]]
    saveName = "红黄蓝亲子园客户导入模板-py生成.xlsx"
    for i in range(numberChosen, 30000 + 1):
        print('正在执行：%d' %i)
        sheet1.cell(row=i, column=1).value = commonUtils.getChineseName()
        sheet1.cell(row=i, column=2).value = "导入测试宝宝"
        sheet1.cell(row=i, column=3).value = "男"
        sheet1.cell(row=i, column=4).value = "20190803"
        sheet1.cell(row=i, column=5).value = commonUtils.getChineseName()
        sheet1.cell(row=i, column=6).value = commonUtils.getNewPhone()
        sheet1.cell(row=i, column=7).value = "妈妈"
        sheet1.cell(row=i, column=8).value = "决策人"
        sheet1.cell(row=i, column=9).value = "互联网"
        sheet1.cell(row=i, column=10).value = "400呼入"
        sheet1.cell(row=i, column=12).value = "叶勇"
        sheet1.cell(row=i, column=14).value = "一般"
    worksheet.save(saveName)
    print('执行完毕')

if __name__ == '__main__':
    main()
