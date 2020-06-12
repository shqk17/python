import pymysql, datetime, lxml
import time
from openpyxl import load_workbook, Workbook
import xlsxwriter


class bsUtil:
    def getDBLink(type):
        if type == 1:
            db = pymysql.connect(host='192.168.0.34',
                                 port=3306,
                                 user='root',
                                 passwd='Aa12345678',
                                 db='tss',
                                 charset='utf8'
                                 , cursorclass=pymysql.cursors.DictCursor
                                 )

        if type == 2:
            db = pymysql.connect(host='hb3-rds20180919.mysql.zhangbei.rds.aliyuncs.com',
                                 port=3382,
                                 user='qms',
                                 passwd='rybu0OsWO2qL7Ef',
                                 db='tss',
                                 charset='utf8'
                                 , cursorclass=pymysql.cursors.DictCursor
                                 )
        if type == 3:
            db = pymysql.connect(host='192.168.0.5',
                                 port=3306,
                                 user='root',
                                 passwd='asdfg_qwert@',
                                 db='tss',
                                 charset='utf8'
                                 , cursorclass=pymysql.cursors.DictCursor
                                 )

        return db.cursor(), db


def getData():
    print("数据库连接中。。。")
    sql = "SELECT c.schoolName, a.`name`, a.sex, a.birthdayDate, a.isMember," \
          " a.stuStatus," \
          " a.createTime, 	count(b.id) as parentNum," \
          " GROUP_CONCAT( CONCAT( b.affiliationRelation,';', b.patriarchName,';', b.contactPhone ) )  AS GX" \
          " FROM" \
          " tss_student a" \
          "	LEFT JOIN tss_pertain_relation b ON a.id = b.studentId" \
          "	LEFT JOIN sys_school c ON a.schoolId = c.id " \
          " WHERE " \
          "	a.isDeleted = 0 and c.schoolName not like '%测试%'" \
          " GROUP BY" \
          "	a.id ORDER BY c.schoolName"

    cursor, db = bsUtil.getDBLink(2)
    print("数据获取中预计耗时1分钟，请耐心等待。。。")
    cursor.execute(sql)
    contents = cursor.fetchall()
    db.close()
    return contents


def getStatus(status):
    try:
        print("当前跟进状态是：" + str(status))
        if status is None or status == '':
            return ""
        status_maps = {"1": "A意向强烈",
                       "2": "B有意向",
                       "0": "",
                       "3": "C一般",
                       "4": "订金",
                       "5": "无效",
                       "6": "历史",
                       "7": "已成交"
                       }
        print("解析后跟进状态是：" + str(status_maps[str(status)]))
        return status_maps[str(status)]
    except Exception as ee:
        print("解析跟进状态异常" + str(ee))
        return " "


def getGx(gx):
    if gx is None or gx == '':
        return ""

    Gx_map = {"1": "妈妈",
              "2": "爸爸",
              "3": "爷爷",
              "4": "奶奶",
              "5": "姥姥",
              "6": "姥爷",
              "7": "司机",
              "8": "保姆",
              "9": "其他",
              "0": " ",
              }

    parents = str(gx).split(",")
    d = ''
    try:
        for p in parents:
            print(str(p))
            if p is None:
                continue
            d = d + "[" + Gx_map[p.split(";")[0]] + "," + p.split(";")[1] + "," + p.split(";")[2] + "] "
        return d
    except Exception as ee:
        print(str(ee))
        return " "


def getSex(sex):
    if sex == "1":
        return "男"
    else:
        return "女"


def getIsMember(isMember):
    if isMember == "1":
        return "是"
    else:
        return "否"


def excel2007xlsx(name, data):
    excel_QA = Workbook()  # 建立一个工作本
    sheet1 = excel_QA.active  # 激活sheet
    sheet1.title = '学生信息导出'  # 对sheet进行命名
    sheet1.cell(row=1, column=1).value = "园所名称"
    sheet1.cell(row=1, column=2).value = "学生姓名"
    sheet1.cell(row=1, column=3).value = "性别"
    sheet1.cell(row=1, column=4).value = "生日"
    sheet1.cell(row=1, column=5).value = "是否是会员"
    sheet1.cell(row=1, column=6).value = "学生状态"
    sheet1.cell(row=1, column=7).value = "创建时间"
    sheet1.cell(row=1, column=8).value = "关系人手机号"
    j = 2
    for i in data:
        j += 1
        print('正在执行：%d' % j)
        try:
            sheet1.cell(row=j, column=1).value = i["schoolName"]
            sheet1.cell(row=j, column=2).value = str(i["name"])
            sheet1.cell(row=j, column=3).value = getSex(i["sex"])
            sheet1.cell(row=j, column=4).value = str(i["birthdayDate"]).split(" ")[0]
            sheet1.cell(row=j, column=5).value = getIsMember(i["isMember"])
            sheet1.cell(row=j, column=6).value = getStatus(i["stuStatus"])
            sheet1.cell(row=j, column=7).value = i["createTime"]
            sheet1.cell(row=j, column=8).value = getGx(i["GX"])
        except Exception as ee:
            print(str(ee))
            continue
    print('共获取'+str(j-2)+'条数据。。。')
    print('数据保存中，预计1~3分钟，请耐心等待。。。')
    excel_QA.save(name)
    print('执行完毕，已保存到同级目录下<qms学生信息导出>。。。')
    input("请输入任意字符,以关闭程序")


def product(contents):
    if len(contents) < 1:
        print("没有查询到数据")
    else:
        print("一共" + str(len(contents)) + "条数据")
        now_time = time.strftime("%Y%m%d%H%M%S", time.localtime())
        excel2007xlsx("qms学生信息导出_"+str(now_time)+".xlsx", contents)


def main():
    product(getData())


if __name__ == '__main__':
    main()
